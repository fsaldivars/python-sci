#
# Alan Badillo Salas <badillo.soft@hotmail.com>
# https://github.com/badillosoft
#
# Python's Scientist Module
#

from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string

def clean_init_spaces(s):
	aux = ""
	i = 0
	while i < len(s) and s[i] == " ":
		i += 1
	while i < len(s):
		aux += s[i]
		i += 1
	return aux

def clean_end_spaces(s):
	aux = ""
	i = 0
	while i < len(s) and s[i] != " " and s[i] != "\n":
		aux += s[i]
		i += 1
	return aux

def range_xl(filename, sheet_name, cell_range):
    wb = load_workbook(filename, data_only=True)
    ws = wb[sheet_name]
    cells = ws[cell_range]
    return cells


def matrix_xl(cells):
    mat = []
    for row_xl in cells:
        row = []
        for cell in row_xl:
            row.append(cell.value)
        mat.append(row)
    return mat


def matrix_csv(filename):
	f = open(filename, "r")
	mat = []
	for line in f:
		row_csv = line.split(",")
		row = []
		for cell in row_csv:
			try:
				row.append(float(cell))
			except:
				s = clean_init_spaces(cell)
				s = clean_end_spaces(s)
				row.append(s)
		mat.append(row)
	f.close()
	return mat


def data_build(mat):
    keys = mat[0]
    coll = []
    for i in range(1, len(mat)):
        dic = {}
        for j in range(0, len(keys)):
			k = keys[j]
			if k != None and k != "":
				dic[k] = mat[i][j]
        coll.append(dic)
    return coll

def load_xl(filename, sheet_name, cell_range):
	cells = range_xl(filename, sheet_name, cell_range)
	mat = matrix_xl(cells)
	return data_build(mat)

def load_csv(filename):
	mat = matrix_csv(filename)
	return data_build(mat)

def data_extract(data, key):
	aux = []
	for dic in data:
		aux.append(dic[key])
	return aux

def data_filter(data, fn_filter):
	aux = []
	for dic in data:
		if fn_filter(dic):
			aux.append(dic)
	return aux

def data_checker(data, fn_checker):
	aux = []
	for dic in data:
		aux.append(fn_checker(dic))
	return aux

def data_reduce(data, reduce):
	aux = []
	for i in range(min(len(data), len(reduce))):
		if (reduce[i]):
			aux.append(data[i])
	return aux

def data_map(data, fn_map):
	aux = []
	for dic in data:
		x = fn_map(dic)
		if x != None:
			aux.append(x)
	return aux

def write_xl(filename, sheet_name, ini_cell, data, labels):
	try:
		wb = load_workbook(filename)
	except:
		wb = Workbook()

	if wb.sheetnames.count(sheet_name) > 0:
		ws = wb[sheet_name]
	else:
		ws = wb.create_sheet(title=sheet_name)
	for j in range(len(labels)):
		cell = ws.cell(
			row = ws[ini_cell].row,
			column = column_index_from_string(ws[ini_cell].column) + j
		)
		cell.value = labels[j]
	for i in range(len(data)):
		dic = data[i]
		for j in range(len(labels)):
			cell = ws.cell(
				row = ws[ini_cell].row + 1 + i,
				column = column_index_from_string(ws[ini_cell].column) + j
			)
			cell.value = dic[labels[j]]
	wb.save(filename)

def cat_build(data, labels):
	return [data_map(data, lambda x: x[k]) for k in labels]

from sets import Set

def cat_set_build(data, labels):
	return [list(set(data_map(data, lambda x: x[k]))) for k in labels]

def cat_row_build(cats, row):
	return [cat[row] for cat in cats]

def cat_join(data, labels):
	cats = cat_build(data, labels)
	return [cat_row_build(cats, i) for i in range(len(data))]

def cat_transform(cat_data, cat_sets):
	new_cat_data = []
	for cat in cat_data:
		new_cat = []
		for i in range(len(cat)):
			new_cat.append(cat_sets[i].index(cat[i]))
		new_cat_data.append(new_cat)
	return new_cat_data

def data_analize(data, xlabels, ylabels):
	X = cat_join(data, xlabels)
	Y = cat_join(data, ylabels)
	return X, Y

def extract_attributes(content, i):
	text = ""
	while i < len(content):
		c = content[i]

		if c == "<" or c == ">":
			break

		text += c
		i += 1

	return [text], i

def extract_name(content, i):
	name = ""
	while i < len(content):
		c = content[i]

		if c == "<" or c == " " or c == ">":
			break

		name += c
		i += 1
	return name, i

def extract_tag(content, i, text):
	state = "name"
	i += 1
	name = ""
	data = ""
	attributes = []
	while i < len(content):
		c = content[i]

		if c == ">":
			i += 1
			state = "correct" if state != "name" else "error"
			break
		if c == "/" and state != "name":
			state = "close"
			i += 2
			break
		if c == "<":
			state = "incomplete"
			break

		if state == "name":
			name, i = extract_name(content, i)
			attributes, i = extract_attributes(content, i)
			state = "into"
			continue

		data += c
		i += 1

	tag = {
		"name": name,
		"attributes": attributes,
		"data": data,
		"text": text
	}
	return tag, i, state

def search_tag(tags, name, i):
	oname = name[1:]
	pos = -1
	sub = 0
	while i < len(tags):
		if tags[i]["name"] == oname:
			sub += 1
		if tags[i]["name"] == name:
			if sub == 0:
				pos = i
				break
			sub -= 1
		i += 1
	return pos

def html_tree(tags, i = 0):
	nodes = []
	while i < len(tags):
		name = tags[i]["name"]
		search_name = "/%s" % name
		j = search_tag(tags, search_name, i + 1)
		# print "TAG %s (%d -> %d)" % (name, i, j)
		if j >= 0:
			# print "INTIAL TAG: %s" % name
			# print "END TAG: %s" % search_name
			sub_tags = tags[i + 1: j]
			# print "SUB TAGS: ", sub_tags
			children = html_tree(sub_tags)
			nodes.append({
				"tag": tags[i],
				"close_tag": tags[j],
				"children": children
			})
			i = j
			continue
		if name[0] == "/":
			i += 1
			continue
		nodes.append({
			"tag": tags[i]
		})
		i += 1
	return nodes

def html_tags(content):
	state = None
	tags = []
	data = ""
	pos = 0
	while pos < len(content):
		c = content[pos]
		if c == "<":
			tag, pos, state = extract_tag(content, pos, data)
			tags.append(tag)
			# print "TAG: %s : %s (%d)" %(str(tag), state, pos)
			data = ""
			continue
		data += c
		pos += 1
	return tags
